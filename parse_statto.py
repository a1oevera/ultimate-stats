#!/usr/bin/env python3
"""
parse_statto.py  —  Turn a Statto ultimate-frisbee JSON export into a clean,
multi-tab .xlsx that imports straight into Google Sheets.

Usage
-----
    python parse_statto.py [input.json] [output.xlsx]

Defaults: input=data.json, output=statto_stats.xlsx

What it produces (tabs)
-----------------------
    Read Me        - column definitions + the assumptions/quirks baked into this file
    Player Totals  - one row per player, whole tournament (all cells are live formulas)
    Player by Game - one row per player per game (live formulas)
    Games          - one row per game: reconstructed score, result, wind
    Points          - one row per real point: lineup, who started O/D, scored/conceded
    Lineups        - LONG format, one row per player per point  (pivot engine for +/-)
    Passes         - one row per pass (546 here), UUIDs resolved to names
    Possessions    - one row per possession
    Blocks         - one row per defensive block

Design notes
------------
* The summary tabs (Player Totals, Player by Game, Games) use COUNTIFS/SUMIFS
  referencing the raw tabs, so they RECALCULATE if you filter or edit the raw
  data in Sheets. Nothing is a hardcoded aggregate.
* Booleans are stored as 1/0 so COUNTIFS is unambiguous across Excel/Sheets/LibreOffice.
* Field coordinates are stored as Statto's normalized 0-1 values (0,0 = one corner,
  1,1 = the diagonal corner). No meter/yard conversion is invented; "ThrowDist" is a
  normalized Euclidean distance, clearly labeled, NOT a real-world distance.

Stat definitions used (edit here if your team scores things differently)
------------------------------------------------------------------------
  Goal          receiver of a pass flagged isAssist
  Assist        thrower of a pass flagged isAssist
  2nd Assist    thrower of a pass flagged isSecondaryAssist (the hockey assist)
  D (Block)     a defensiveBlocks event credited to the player
  Throwaway     a pass flagged isThrowerError (credited to the thrower)
  Drop          a pass flagged isReceiverError (credited to the receiver)
  Turnover      Throwaways + Drops
  Throw         any pass thrown by the player
  Completion    a throw that was NOT a throwaway
  Completion%   Completions / Throws
  Reception     receiver on a completed pass (not a throwaway, not a drop)
  Point Played  the player was one of the 7 on the field for that point
  +/-           (points won while on field) - (points lost while on field)
"""

import json
import math
import sys
import zipfile
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------------- styling
FONT_NAME   = "Arial"
HDR_FILL    = PatternFill("solid", fgColor="1F3B57")   # dark slate
HDR_FONT    = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
TITLE_FONT  = Font(name=FONT_NAME, bold=True, size=14, color="1F3B57")
BASE_FONT   = Font(name=FONT_NAME, size=10)
BOLD_FONT   = Font(name=FONT_NAME, size=10, bold=True)
ALT_FILL    = PatternFill("solid", fgColor="EEF3F8")   # zebra
THIN        = Side(style="thin", color="D0D7DE")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _norm_dist(x1, y1, x2, y2):
    if None in (x1, y1, x2, y2):
        return None
    return round(math.hypot(x2 - x1, y2 - y1), 4)


def _uuid_str(v):
    """Statto uses {} instead of a UUID when a field is unset."""
    return v if isinstance(v, str) else None


def load_team(path):
    if zipfile.is_zipfile(path):
        # .statto exports are a zip containing data.json (Statto's native format).
        with zipfile.ZipFile(path) as zf:
            with zf.open("data.json") as fh:
                doc = json.load(fh)
    else:
        with open(path, "r", encoding="utf-8") as fh:
            doc = json.load(fh)
    teams = doc.get("teams", [])
    if not teams:
        raise SystemExit("No teams found in export.")
    if len(teams) > 1:
        print(f"WARNING: export has {len(teams)} teams; using the first "
              f"('{teams[0]['data']['team'].get('name')}').", file=sys.stderr)
    return doc, teams[0]["data"]


def build_tables(team):
    rel = team["relations"]

    players = {p["uuid"]: p for p in rel["players"]}
    def pname(uuid):
        u = _uuid_str(uuid)
        if u is None:
            return "(unattributed)"
        p = players.get(u)
        return p["name"] if p else "(unknown)"

    games = {g["uuid"]: g for g in rel["games"]}
    def gname(uuid):
        g = games.get(uuid)
        return g["opponent"] if g else "(unknown game)"

    # real points only (Statto logs empty result=0 marker points with 0 players)
    points = {p["uuid"]: p for p in rel["points"] if p.get("playerUUIDs")}
    possessions = {p["uuid"]: p for p in rel["possessions"]}

    # ---- per-game score reconstruction (points with result +1 / -1) ----
    score = defaultdict(lambda: [0, 0])   # game_uuid -> [for, against]
    for p in rel["points"]:
        if p["result"] == 1:
            score[p["gameUUID"]][0] += 1
        elif p["result"] == -1:
            score[p["gameUUID"]][1] += 1

    games_rows = []
    for g in sorted(rel["games"], key=lambda x: x["date"]):
        f, a = score[g["uuid"]]
        games_rows.append({
            "Game": g["opponent"],
            "Date": g["date"][:10],
            "Result": {1: "W", -1: "L", 0: "T"}.get(g["result"], "?"),
            "Us": f, "Them": a,
            "Wind (kph)": g.get("windSpeed", ""),
        })

    # ---- points + lineups (long) ----
    points_rows, lineup_rows = [], []
    # stable point numbering per game, in creation order
    pt_order = defaultdict(int)
    pt_number = {}
    for p in sorted(points.values(), key=lambda x: x["createdAt"]):
        pt_order[p["gameUUID"]] += 1
        pt_number[p["uuid"]] = pt_order[p["gameUUID"]]

    # ---- pod (group) assignment per point ----
    # Pods overlap heavily, so a point is INFERRED to the pod sharing the most
    # on-field players; ties break toward the smaller (tighter) pod. Pod Match
    # flags whether all 7 belonged to that pod ('exact') or not ('partial n/7').
    pod_defs = [(g["title"], set(g["playerUUIDs"]), len(g["playerUUIDs"]))
                for g in rel.get("groups", [])]

    def assign_pod(uuids):
        onf = set(uuids)
        if not pod_defs or not onf:
            return "(none)", 0, True
        ranked = sorted(((len(onf & mem), -size, name) for name, mem, size in pod_defs),
                        reverse=True)
        n, _, name = ranked[0]
        return name, n, (n == len(onf))

    point_pod, point_match = {}, {}
    for u, p in points.items():
        name, n, exact = assign_pod(p["playerUUIDs"])
        point_pod[u] = name
        point_match[u] = "exact" if exact else f"partial {n}/{len(p['playerUUIDs'])}"

    for p in sorted(points.values(), key=lambda x: (games[p["gameUUID"]]["date"], pt_number[p["uuid"]])):
        started = "O" if p["isOffense"] else "D"
        result = {1: "Scored", -1: "Conceded", 0: "—"}.get(p["result"], "?")
        pm = {1: 1, -1: -1}.get(p["result"], 0)
        line = [pname(u) for u in p["playerUUIDs"]]
        points_rows.append({
            "Game": gname(p["gameUUID"]),
            "Point": pt_number[p["uuid"]],
            "Started": started,
            "Result": result,
            "Pod": point_pod[p["uuid"]],
            "Pod Match": point_match[p["uuid"]],
            "Line": ", ".join(line),
        })
        for u in p["playerUUIDs"]:
            lineup_rows.append({
                "Game": gname(p["gameUUID"]),
                "Point": pt_number[p["uuid"]],
                "Started": started,
                "Result": result,
                "PlusMinus": pm,
                "Player": pname(u),
                "Pod": point_pod[p["uuid"]],
            })

    # ---- passes ----
    poss_point = {u: po["pointUUID"] for u, po in possessions.items()}
    passes_rows = []
    for pa in rel["passes"]:
        puuid = pa["possessionUUID"]
        ptu = poss_point.get(puuid)
        pt = points.get(ptu) if ptu else None
        is_to = int(pa["isThrowerError"] or pa["isReceiverError"])  # one lost disc, even if both flags set
        passes_rows.append({
            "Game": gname(pt["gameUUID"]) if pt else "(unknown)",
            "Point": pt_number.get(ptu, ""),
            "Thrower": pname(pa["throwerUUID"]),
            "Receiver": pname(pa["receiverUUID"]) if _uuid_str(pa["receiverUUID"]) else "",
            "Assist": int(pa["isAssist"]),
            "2nd Assist": int(pa["isSecondaryAssist"]),
            "Throwaway": int(pa["isThrowerError"]),
            "Drop": int(pa["isReceiverError"]),
            "Is TO": is_to,
            "StartX": round(pa["startX"], 4), "StartY": round(pa["startY"], 4),
            "EndX": round(pa["endX"], 4), "EndY": round(pa["endY"], 4),
            "ThrowDist (norm)": _norm_dist(pa["startX"], pa["startY"], pa["endX"], pa["endY"]),
            "Pod": point_pod.get(ptu, ""),
        })

    # stall-outs against us (a turnover with no throw, so absent from the passes list)
    stallout_rows = []
    for s in rel.get("stallOutsAgainst", []):
        ptu = poss_point.get(s.get("possessionUUID"))
        pt = points.get(ptu) if ptu else None
        stallout_rows.append({
            "Game": gname(pt["gameUUID"]) if pt else "(unknown)",
            "Point": pt_number.get(ptu, ""),
            "Player": pname(s.get("playerUUID")),
            "Pod": point_pod.get(ptu, ""),
        })

    # ---- possessions ----
    passes_by_poss = defaultdict(list)
    for pa in rel["passes"]:
        passes_by_poss[pa["possessionUUID"]].append(pa)
    poss_rows = []
    for po in possessions.values():
        pt = points.get(po["pointUUID"])
        plist = passes_by_poss.get(po["uuid"], [])
        ended_goal = int(any(pa["isAssist"] for pa in plist))
        poss_rows.append({
            "Game": gname(pt["gameUUID"]) if pt else "(unknown)",
            "Point": pt_number.get(po["pointUUID"], ""),
            "Initiator": pname(po["initiatorUUID"]),
            "StartX": round(po["startX"], 4), "StartY": round(po["startY"], 4),
            "Passes": len(plist),
            "Ended in Goal": ended_goal,
            "Pod": point_pod.get(po["pointUUID"], ""),
        })

    # ---- blocks ----
    blocks_rows = []
    for b in rel["defensiveBlocks"]:
        pt = points.get(b["pointUUID"])
        blocks_rows.append({
            "Game": gname(pt["gameUUID"]) if pt else "(unknown)",
            "Point": pt_number.get(b["pointUUID"], ""),
            "Player": pname(b["playerUUID"]),
            "Stall Out": int(b["isStallOut"]),
            "Callahan": int(b["isCallahan"]),
            "LocX": round(b["locationX"], 4), "LocY": round(b["locationY"], 4),
            "Pod": point_pod.get(b["pointUUID"], ""),
        })

    roster = sorted(rel["players"], key=lambda p: (p.get("sortName") or p["name"]).lower())
    roster_rows = [{"Player": p["name"], "Number": p.get("number", "")} for p in roster]
    pod_rows = [{"Pod": name, "Players": size} for name, _, size in pod_defs]

    return {
        "games": games_rows, "points": points_rows, "lineups": lineup_rows,
        "passes": passes_rows, "possessions": poss_rows, "blocks": blocks_rows,
        "stallouts": stallout_rows, "roster": roster_rows, "pods": pod_rows,
        "team_name": team["team"].get("name", "Team"),
    }


# ----------------------------------------------------------------------------- writing
def _write_grid(ws, headers, rows, autofilter=True, freeze="A2"):
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    for i, r in enumerate(rows, start=2):
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=i, column=c, value=r.get(h))
            cell.font = BASE_FONT
            cell.border = BORDER
            if i % 2 == 0:
                cell.fill = ALT_FILL
    if freeze:
        ws.freeze_panes = freeze
    if autofilter and rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    _autosize(ws, headers, rows)


def _autosize(ws, headers, rows):
    for c, h in enumerate(headers, start=1):
        width = len(str(h))
        for r in rows:
            v = r.get(h)
            if v is not None:
                width = max(width, len(str(v)))
        ws.column_dimensions[get_column_letter(c)].width = min(max(width + 2, 8), 48)


def build_workbook(t):
    wb = Workbook()

    # ---------- Read Me ----------
    ws = wb.active
    ws.title = "Read Me"
    ws["A1"] = f"{t['team_name']} — Statto tournament stats"
    ws["A1"].font = TITLE_FONT
    legend = [
        "", "TABS",
        "  Player Totals  – per player, whole tournament (live formulas)",
        "  Player by Game – per player per game (live formulas)",
        "  Pods           – per-pod stats (holds, breaks, clean%, TOs, goals, Ds)",
        "  Games          – reconstructed score, result, wind",
        "  Points         – one row per point: O/D start, scored/conceded, Our TOs, Clean, Pod, lineup",
        "  Lineups        – LONG: one row per player per point (pivot this for +/-)",
        "  Passes         – one row per pass, names resolved from UUIDs",
        "  Possessions    – one row per possession",
        "  Blocks         – one row per defensive block (Ds our team got)",
        "  StallOuts      – stall-outs AGAINST us (a turnover with no throw)",
        "  (every event tab carries a 'Pod' column so you can pivot any stat by pod)",
        "",
        "STAT DEFINITIONS",
        "  Goal        = receiver of an assist pass",
        "  Assist      = thrower of an assist pass",
        "  2nd Assist  = thrower of the secondary-assist pass (hockey assist)",
        "  D           = defensive block credited to the player",
        "  Throwaway   = pass flagged as a thrower error",
        "  Drop        = pass flagged as a receiver error",
        "  Turnover    = Throwaways + Drops",
        "  Completion% = (throws that were not throwaways) / all throws",
        "  Reception   = receiver on a completed, non-dropped pass",
        "  +/-         = points won on field minus points lost on field",
        "",
        "CLEAN POINTS (on the Points tab)",
        "  Our TOs = turnovers by our team on that point = passes flagged 'Is TO'",
        "            (a throwaway OR a drop, counted once) + stall-outs against us.",
        "  Clean   = 'Y' when Our TOs = 0, else 'N'.",
        "  Clean hold  = Started 'O'  AND Result 'Scored'   AND Clean = 'Y'",
        "  Clean break = Started 'D'  AND Result 'Scored'   AND Clean = 'Y'",
        "  A point can be Clean='Y' but conceded (Started 'D', we never won the disc back).",
        "  Note: 'Is TO' counts one lost disc even for the ~10 passes Statto flags as",
        "  BOTH a throwaway and a drop, so Our TOs is not Throwaways+Drops summed.",
        "",
        "ASSUMPTIONS & DATA QUIRKS (verify against your own scoring)",
        "  • Empty 'marker' points (Statto result=0, 0 players) are excluded from",
        "    Points / Lineups. They are game/half boundaries, not real points.",
        "  • Events Statto left unattributed (no player UUID) show as '(unattributed)'.",
        "    They appear in raw tabs and team counts but are NOT credited to a player.",
        "  • Field coordinates are Statto's NORMALIZED 0–1 values. 'ThrowDist (norm)'",
        "    is a normalized Euclidean distance, not meters/yards.",
        "  • Booleans are stored as 1/0 so COUNTIFS behaves consistently.",
        "  • This export contains one team only — there are no opponent player stats.",
        "  • POD tags are INFERRED, not recorded. Statto stores pod rosters, not which",
        "    pod played each point. A point is tagged to the pod sharing the most on-field",
        "    players; ties break toward the smaller pod. 'Pod Match' on the Points tab",
        "    shows 'exact' (all 7 in that pod) or 'partial n/7' — 88 exact, 4 partial here.",
        "    Pods overlap heavily: Pod patrol and Power share exactly 7 players, so lineups",
        "    of those 7 are ambiguous and default to Pod patrol. Treat pod stats as",
        "    approximate, and re-tag the 'partial' points by hand if you know the truth.",
    ]
    for i, line in enumerate(legend, start=2):
        ws.cell(row=i, column=1, value=line).font = (
            BOLD_FONT if line.strip() in
            ("TABS", "STAT DEFINITIONS", "ASSUMPTIONS & DATA QUIRKS (verify against your own scoring)")
            else BASE_FONT
        )
    ws.column_dimensions["A"].width = 78

    # ---------- raw tabs ----------
    passes_h = ["Game", "Point", "Thrower", "Receiver", "Assist", "2nd Assist",
                "Throwaway", "Drop", "Is TO", "StartX", "StartY", "EndX", "EndY",
                "ThrowDist (norm)", "Pod"]
    lineup_h = ["Game", "Point", "Started", "Result", "PlusMinus", "Player", "Pod"]
    poss_h   = ["Game", "Point", "Initiator", "StartX", "StartY", "Passes", "Ended in Goal", "Pod"]
    blocks_h = ["Game", "Point", "Player", "Stall Out", "Callahan", "LocX", "LocY", "Pod"]
    stall_h  = ["Game", "Point", "Player", "Pod"]

    _write_grid(wb.create_sheet("Games"),
                ["Game", "Date", "Result", "Us", "Them", "Wind (kph)"], t["games"])
    ws_line   = wb.create_sheet("Lineups"); _write_grid(ws_line, lineup_h, t["lineups"])
    ws_pass   = wb.create_sheet("Passes"); _write_grid(ws_pass, passes_h, t["passes"])
    _write_grid(wb.create_sheet("Possessions"), poss_h, t["possessions"])
    ws_block  = wb.create_sheet("Blocks"); _write_grid(ws_block, blocks_h, t["blocks"])
    ws_stall  = wb.create_sheet("StallOuts"); _write_grid(ws_stall, stall_h, t["stallouts"])

    # ---------- Points (with live turnover / clean flags) ----------
    NPa = len(t["passes"]) + 1
    NSo = len(t["stallouts"]) + 1
    # Passes cols: A Game B Point ... I Is TO ; StallOuts cols: A Game B Point
    def pg(c): return f"Passes!${c}$2:${c}${NPa}"
    def sg(c): return f"StallOuts!${c}$2:${c}${NSo}"
    ws_points = wb.create_sheet("Points")
    points_h = ["Game", "Point", "Started", "Result", "Our TOs", "Clean", "Pod", "Pod Match", "Line"]
    ws_points.append(points_h)
    for c in range(1, len(points_h) + 1):
        cell = ws_points.cell(row=1, column=c)
        cell.font, cell.fill = HDR_FONT, HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    for i, r in enumerate(t["points"], start=2):
        ws_points.cell(row=i, column=1, value=r["Game"])
        ws_points.cell(row=i, column=2, value=r["Point"])
        ws_points.cell(row=i, column=3, value=r["Started"])
        ws_points.cell(row=i, column=4, value=r["Result"])
        # turnovers by us on this point = passes flagged Is TO + stall-outs against, keyed on Game+Point
        to_pass = f'COUNTIFS({pg("A")},$A{i},{pg("B")},$B{i},{pg("I")},1)'
        to_stall = (f'+COUNTIFS({sg("A")},$A{i},{sg("B")},$B{i})'
                    if len(t["stallouts"]) else "")
        ws_points.cell(row=i, column=5, value=f'={to_pass}{to_stall}')
        ws_points.cell(row=i, column=6, value=f'=IF(E{i}=0,"Y","N")')
        ws_points.cell(row=i, column=7, value=r["Pod"])
        ws_points.cell(row=i, column=8, value=r["Pod Match"])
        ws_points.cell(row=i, column=9, value=r["Line"])
        for c in range(1, len(points_h) + 1):
            cell = ws_points.cell(row=i, column=c)
            cell.font = BASE_FONT; cell.border = BORDER
            if i % 2 == 0:
                cell.fill = ALT_FILL
    ws_points.freeze_panes = "A2"
    ws_points.auto_filter.ref = f"A1:I{len(t['points']) + 1}"
    for c, w in zip(range(1, 10), [12, 6, 8, 10, 8, 7, 12, 12, 60]):
        ws_points.column_dimensions[get_column_letter(c)].width = w

    # row extents for formula ranges
    NP = len(t["passes"]) + 1
    NL = len(t["lineups"]) + 1
    NB = len(t["blocks"]) + 1
    # column letters on Passes: A Game B Point C Thrower D Receiver E Assist F 2nd G Throwaway H Drop
    P = {"Game": "A", "Thrower": "C", "Receiver": "D", "Assist": "E",
         "Sec": "F", "Throwaway": "G", "Drop": "H"}
    # Lineups: A Game B Point C Started D Result E PlusMinus F Player
    L = {"Game": "A", "PM": "E", "Player": "F"}
    # Blocks: A Game B Point C Player
    B = {"Game": "A", "Player": "C"}

    def prng(col):  return f"Passes!${col}$2:${col}${NP}"
    def lrng(col):  return f"Lineups!${col}$2:${col}${NL}"
    def brng(col):  return f"Blocks!${col}$2:${col}${NB}"

    # ---------- Player by Game ----------
    wsg = wb.create_sheet("Player by Game")
    pbg_headers = ["Player", "Game", "GP", "Pts", "+/-", "G", "A", "2A", "D", "TA", "Drops"]
    wsg.append(pbg_headers)
    for c in range(1, len(pbg_headers) + 1):
        cell = wsg.cell(row=1, column=c)
        cell.font, cell.fill = HDR_FONT, HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    game_labels = [g["Game"] for g in t["games"]]
    r = 2
    for pr in t["roster"]:
        name = pr["Player"]
        for gl in game_labels:
            A = f'"{name}"'; G = f'"{gl}"'
            pts = f'COUNTIFS({lrng(L["Player"])},{A},{lrng(L["Game"])},{G})'
            wsg.cell(row=r, column=1, value=name)
            wsg.cell(row=r, column=2, value=gl)
            wsg.cell(row=r, column=3, value=f'=IF({pts}>0,1,0)')
            wsg.cell(row=r, column=4, value=f'={pts}')
            wsg.cell(row=r, column=5, value=f'=SUMIFS({lrng(L["PM"])},{lrng(L["Player"])},{A},{lrng(L["Game"])},{G})')
            wsg.cell(row=r, column=6, value=f'=COUNTIFS({prng(P["Receiver"])},{A},{prng(P["Assist"])},1,{prng(P["Game"])},{G})')
            wsg.cell(row=r, column=7, value=f'=COUNTIFS({prng(P["Thrower"])},{A},{prng(P["Assist"])},1,{prng(P["Game"])},{G})')
            wsg.cell(row=r, column=8, value=f'=COUNTIFS({prng(P["Thrower"])},{A},{prng(P["Sec"])},1,{prng(P["Game"])},{G})')
            wsg.cell(row=r, column=9, value=f'=COUNTIFS({brng(B["Player"])},{A},{brng(B["Game"])},{G})')
            wsg.cell(row=r, column=10, value=f'=COUNTIFS({prng(P["Thrower"])},{A},{prng(P["Throwaway"])},1,{prng(P["Game"])},{G})')
            wsg.cell(row=r, column=11, value=f'=COUNTIFS({prng(P["Receiver"])},{A},{prng(P["Drop"])},1,{prng(P["Game"])},{G})')
            for c in range(1, len(pbg_headers) + 1):
                cell = wsg.cell(row=r, column=c)
                cell.font = BASE_FONT; cell.border = BORDER
                if r % 2 == 0:
                    cell.fill = ALT_FILL
            r += 1
    wsg.freeze_panes = "A2"
    wsg.auto_filter.ref = f"A1:K{r-1}"
    for c, h in enumerate(pbg_headers, 1):
        wsg.column_dimensions[get_column_letter(c)].width = max(len(h) + 2, 8)
    PBG_LAST = r - 1

    # ---------- Player Totals ----------
    wst = wb.create_sheet("Player Totals")
    tot_headers = ["Player", "#", "GP", "Pts", "+/-", "G", "A", "2A", "G+A", "D",
                   "TA", "Drops", "TO", "Throws", "Comp", "Comp%", "Rec"]
    wst.append(tot_headers)
    for c in range(1, len(tot_headers) + 1):
        cell = wst.cell(row=1, column=c)
        cell.font, cell.fill = HDR_FONT, HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    # Player-by-Game columns: A Player B Game C GP ...
    def gp_sumifs(name):
        return f'SUMIFS(\'Player by Game\'!$C$2:$C${PBG_LAST},\'Player by Game\'!$A$2:$A${PBG_LAST},"{name}")'

    r = 2
    for pr in t["roster"]:
        name = pr["Player"]; A = f'"{name}"'
        throws = f'COUNTIF({prng(P["Thrower"])},{A})'
        comp   = f'COUNTIFS({prng(P["Thrower"])},{A},{prng(P["Throwaway"])},0)'
        wst.cell(row=r, column=1, value=name)
        wst.cell(row=r, column=2, value=pr["Number"])
        wst.cell(row=r, column=3, value=f'={gp_sumifs(name)}')
        wst.cell(row=r, column=4, value=f'=COUNTIF({lrng(L["Player"])},{A})')
        wst.cell(row=r, column=5, value=f'=SUMIFS({lrng(L["PM"])},{lrng(L["Player"])},{A})')
        wst.cell(row=r, column=6, value=f'=COUNTIFS({prng(P["Receiver"])},{A},{prng(P["Assist"])},1)')
        wst.cell(row=r, column=7, value=f'=COUNTIFS({prng(P["Thrower"])},{A},{prng(P["Assist"])},1)')
        wst.cell(row=r, column=8, value=f'=COUNTIFS({prng(P["Thrower"])},{A},{prng(P["Sec"])},1)')
        wst.cell(row=r, column=9, value=f'=F{r}+G{r}')
        wst.cell(row=r, column=10, value=f'=COUNTIF({brng(B["Player"])},{A})')
        wst.cell(row=r, column=11, value=f'=COUNTIFS({prng(P["Thrower"])},{A},{prng(P["Throwaway"])},1)')
        wst.cell(row=r, column=12, value=f'=COUNTIFS({prng(P["Receiver"])},{A},{prng(P["Drop"])},1)')
        wst.cell(row=r, column=13, value=f'=K{r}+L{r}')
        wst.cell(row=r, column=14, value=f'={throws}')
        wst.cell(row=r, column=15, value=f'={comp}')
        wst.cell(row=r, column=16, value=f'=IFERROR(O{r}/N{r},0)')
        wst.cell(row=r, column=17, value=f'=COUNTIFS({prng(P["Receiver"])},{A},{prng(P["Throwaway"])},0,{prng(P["Drop"])},0)')
        for c in range(1, len(tot_headers) + 1):
            cell = wst.cell(row=r, column=c)
            cell.font = BASE_FONT; cell.border = BORDER
            if r % 2 == 0:
                cell.fill = ALT_FILL
        wst.cell(row=r, column=16).number_format = "0.0%"
        r += 1
    wst.freeze_panes = "C2"
    wst.auto_filter.ref = f"A1:{get_column_letter(len(tot_headers))}{r-1}"
    widths = [18, 4] + [6] * (len(tot_headers) - 2)
    for c, w in enumerate(widths, 1):
        wst.column_dimensions[get_column_letter(c)].width = w
    wst.column_dimensions["P"].width = 7

    # ---------- Pods ----------
    # Points cols: A Game B Point C Started D Result E Our TOs F Clean G Pod H Pod Match I Line
    # Passes Pod col = O (15th) ; Blocks Pod col = H (8th)
    NPt = len(t["points"]) + 1
    NPa2 = len(t["passes"]) + 1
    NBl = len(t["blocks"]) + 1
    ptG = f"Points!$G$2:$G${NPt}"      # Pod
    ptC = f"Points!$C$2:$C${NPt}"      # Started
    ptD = f"Points!$D$2:$D${NPt}"      # Result
    ptE = f"Points!$E$2:$E${NPt}"      # Our TOs
    ptF = f"Points!$F$2:$F${NPt}"      # Clean
    paPod = f"Passes!$O$2:$O${NPa2}"
    paA   = f"Passes!$E$2:$E${NPa2}"   # Assist
    blPod = f"Blocks!$H$2:$H${NBl}"

    wsp = wb.create_sheet("Pods")
    pod_headers = ["Pod", "Players", "Points", "O Pts", "D Pts", "Scored", "Conceded",
                   "Record", "Holds", "Hold%", "Breaks", "Break%", "Clean Holds",
                   "Clean Breaks", "Our TOs", "TOs/Pt", "Goals", "Ds"]
    wsp.append(pod_headers)
    for c in range(1, len(pod_headers) + 1):
        cell = wsp.cell(row=1, column=c)
        cell.font, cell.fill = HDR_FONT, HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    r = 2
    for pod in t["pods"]:
        A = f'"{pod["Pod"]}"'
        wsp.cell(row=r, column=1, value=pod["Pod"])
        wsp.cell(row=r, column=2, value=pod["Players"])
        wsp.cell(row=r, column=3, value=f'=COUNTIF({ptG},{A})')
        wsp.cell(row=r, column=4, value=f'=COUNTIFS({ptG},{A},{ptC},"O")')
        wsp.cell(row=r, column=5, value=f'=COUNTIFS({ptG},{A},{ptC},"D")')
        wsp.cell(row=r, column=6, value=f'=COUNTIFS({ptG},{A},{ptD},"Scored")')
        wsp.cell(row=r, column=7, value=f'=COUNTIFS({ptG},{A},{ptD},"Conceded")')
        wsp.cell(row=r, column=8, value=f'=F{r}&"-"&G{r}')
        wsp.cell(row=r, column=9, value=f'=COUNTIFS({ptG},{A},{ptC},"O",{ptD},"Scored")')
        wsp.cell(row=r, column=10, value=f'=IFERROR(I{r}/D{r},0)')
        wsp.cell(row=r, column=11, value=f'=COUNTIFS({ptG},{A},{ptC},"D",{ptD},"Scored")')
        wsp.cell(row=r, column=12, value=f'=IFERROR(K{r}/E{r},0)')
        wsp.cell(row=r, column=13, value=f'=COUNTIFS({ptG},{A},{ptC},"O",{ptD},"Scored",{ptF},"Y")')
        wsp.cell(row=r, column=14, value=f'=COUNTIFS({ptG},{A},{ptC},"D",{ptD},"Scored",{ptF},"Y")')
        wsp.cell(row=r, column=15, value=f'=SUMIFS({ptE},{ptG},{A})')
        wsp.cell(row=r, column=16, value=f'=IFERROR(O{r}/C{r},0)')
        wsp.cell(row=r, column=17, value=f'=COUNTIFS({paPod},{A},{paA},1)')
        wsp.cell(row=r, column=18, value=f'=COUNTIF({blPod},{A})')
        for c in range(1, len(pod_headers) + 1):
            cell = wsp.cell(row=r, column=c)
            cell.font = BASE_FONT; cell.border = BORDER
            if r % 2 == 0:
                cell.fill = ALT_FILL
        wsp.cell(row=r, column=10).number_format = "0.0%"
        wsp.cell(row=r, column=12).number_format = "0.0%"
        wsp.cell(row=r, column=16).number_format = "0.00"
        r += 1
    wsp.freeze_panes = "B2"
    for c, h in enumerate(pod_headers, 1):
        wsp.column_dimensions[get_column_letter(c)].width = max(len(h) + 1, 7)
    wsp.column_dimensions["A"].width = 14

    desired = ["Read Me", "Player Totals", "Player by Game", "Pods", "Games",
               "Points", "Lineups", "Passes", "Possessions", "Blocks", "StallOuts"]
    wb._sheets.sort(key=lambda s: desired.index(s.title) if s.title in desired else 99)
    return wb


def main():
    inp = sys.argv[1] if len(sys.argv) > 1 else "data.json"
    out = sys.argv[2] if len(sys.argv) > 2 else "statto_stats.xlsx"
    _, team = load_team(inp)
    tables = build_tables(team)
    wb = build_workbook(tables)
    wb.save(out)
    print(f"Wrote {out}  ({len(tables['roster'])} players, "
          f"{len(tables['games'])} games, {len(tables['passes'])} passes)")


if __name__ == "__main__":
    main()